import operator
from random import randint

import openpyxl.styles


def fifo(pages, amountOfFrames):
    pageFaults = 0
    frames = []
    stringOfFrames = []
    for page in pages:
        if page not in frames:
            pageFaults += 1
            if len(frames) < amountOfFrames:
                frames.append(page)
            else:
                del frames[0]
                frames.append(page)
        stringOfFrames.append(frames[:])
    return stringOfFrames, pageFaults


def lru(pages, amountOfFrames):
    # zdefiniowanie błedów ramki, ramek (słownik) i licznika- do określenia, kiedy zostało wysłane żądanie
    stringOfFrames = []
    pageFaults = 0
    frames = {}  # klucze- numery stron, wartość - ostatnie odwołania
    counter = 1
    for num in pages:
        # sprawdzenie czy danej strony nie mana ramce
        if num not in frames.keys():
            pageFaults += 1
            # sprawdzenie liczebności ramek- jeśli nie są zapełnione, po prostu dodajemy
            if len(frames.keys()) < amountOfFrames:
                frames[num] = counter
            # jeśli wszystkie są zajęte- trzeba wybrać ofiarę- ta, która była używana najdawniej
            else:
                del frames[min(frames.items(), key=operator.itemgetter(1))[0]]
                frames[num] = counter
        else:
            frames[num] = counter
        stringOfFrames.append(frames.copy())
        counter += 1
    return stringOfFrames, pageFaults


def lfu(pages, amountOfFrames):
    # zdefiniowanie błedów ramki, ramek (słownik) i licznika- do określenia, kiedy zostało wysłane żądanie
    pageFaults = 0
    stringOfFrames = []
    frames = {}  # klucze- numery stron, wartość - ilość odwołań
    references = {}  # przechwoywana tu jest ilość odwołań
    for num in pages:
        # sprawdzenie czy danej strony nie ma na ramce
        if num not in frames.keys():
            pageFaults += 1
            # sprawdzenie liczebności ramek- jeśli nie są zapełnione
            if len(frames.keys()) < amountOfFrames:
                # sprawdzenie czy do tego numeru było już wcześniej jakieś odwołanie- jeśli nie dodajemy 1, jeśli tak, do ilości odwołań dodajemy 1
                if num not in references.keys():
                    frames[num] = 1
                else:
                    frames[num] = references[num] + 1
            # jeśli wszystkie ramki są zapełnione- trzeba wybrać ofiarę- tą najczęściej używaną- z największą value
            else:
                # dodanie wartości do słownika spisującego
                idxMFU = min(frames.items(), key=operator.itemgetter(1))[0]
                if idxMFU in references.keys():
                    references[idxMFU] += 1
                else:
                    references[idxMFU] = frames[idxMFU]
                del frames[idxMFU]
                # sprawdzam czy odwołanie wcześniej miało już miejsce
                if num in references:
                    frames[num] = references[num] + 1
                else:
                    frames[num] = 1
        else:
            tmp = frames[num]
            del frames[num]
            frames[num] = tmp + 1
        stringOfFrames.append(frames.copy())
    return stringOfFrames, pageFaults


def mfu(pages, amountOfFrames):
    # zdefiniowanie błedów ramki, ramek (słownik) i licznika- do określenia, kiedy zostało wysłane żądanie
    stringOfFrames = []
    pageFaults = 0
    frames = {}  # klucze- numery stron, wartość - ilość odwołań
    references = {}  # przechwoywana tu jest ilość odwołań
    for num in pages:
        # sprawdzenie czy danej strony nie ma na ramce
        if num not in frames.keys():
            pageFaults += 1
            # sprawdzenie liczebności ramek- jeśli nie są zapełnione
            if len(frames.keys()) < amountOfFrames:
                # sprawdzenie czy do tego numeru było już wcześniej jakieś odwołanie- jeśli nie dodajemy 1, jeśli tak, do ilości odwołań dodajemy 1
                if num not in references.keys():
                    frames[num] = 1
                    # references[num] = 1
                else:
                    frames[num] = references[num] + 1
            # jeśli wszystkie ramki są zapełnione- trzeba wybrać ofiarę- tą najczęściej używaną- z największą value
            else:
                # dodanie wartości do słownika spisującego
                idxMFU = max(frames.items(), key=operator.itemgetter(1))[0]
                if idxMFU in references.keys():
                    references[idxMFU] += 1
                else:
                    references[idxMFU] = frames[idxMFU]
                del frames[idxMFU]
                # sprawdzam czy odwołanie wcześniej miało już miejsce
                if num in references:
                    frames[num] = references[num] + 1
                else:
                    frames[num] = 1
        else:
            tmp = frames[num]
            del frames[num]
            frames[num] = tmp + 1
        stringOfFrames.append(frames.copy())
    return stringOfFrames, pageFaults


def printFrames(frames):
    if isinstance(frames, dict):
        for el in frames.keys():
            print(el, '(' + str(frames[el]) + ') ', end=' ')
        print()
    else:
        for el in frames:
            for k in el:
                print(k, end=' ')
            print()


def BeladyAnomaly(firstNumOfPG, secondNumOfPG, start, stop, howMuch):
    i = 0

    if firstNumOfPG > secondNumOfPG:
        secondNumOfPG, firstNumOfPG = firstNumOfPG, secondNumOfPG
    elif firstNumOfPG == secondNumOfPG:
        return -1
    while True:
        if i > 1500000: break

        pages = [randint(start, stop) for i in range(howMuch)]
        fifoL, faultsFIFOL = fifo(pages, firstNumOfPG)
        fifo4, faultsFIFOM = fifo(pages, secondNumOfPG)
        lrul, faultsLRUL = lru(pages, firstNumOfPG)
        lrum, faultsLRUM = lru(pages, secondNumOfPG)
        mfuL, faultsMFUL = mfu(pages, firstNumOfPG)
        mfuM, faultsMFUM = mfu(pages, secondNumOfPG)
        lfuL, faultsLFUL = lfu(pages, firstNumOfPG)
        lfuM, faultsLFUM = lfu(pages, secondNumOfPG)

        if faultsLRUL < faultsLRUM:
            return lrul, faultsLRUL, lrum, faultsLRUM
        if faultsLFUL < faultsLFUM:
            return lfuL, faultsLFUL, lfuM, faultsLFUM
        if faultsMFUL < faultsMFUM:
            return mfuL, faultsMFUL, mfuM, faultsMFUM
        if faultsFIFOL < faultsFIFOM:
            return fifoL, faultsFIFOL, fifo4, faultsFIFOM
        i += 1


import openpyxl  as opxl
import os
from string import ascii_uppercase

os.chdir('C:\\Users\\user\\PycharmProjects\\projektSO\\')

import json
def exportList(pages, fileName):
    with open(fileName, 'w') as filehandle:
        json.dump(pages, filehandle)



def importList(fileName):
    with open(fileName, 'r') as filehandle:
        pages = json.load(filehandle)
    return pages

def importData(frames, amountOfFrames, pageFaults, name, fileName):
    file = opxl.Workbook()
    sheet = file.create_sheet(name)
    sheet['A1'] = 'RAMKI'
    sheet[str(ascii_uppercase[amountOfFrames + 5]) + str(1)] = 'Błędów braku strony: '
    sheet[str(ascii_uppercase[amountOfFrames + 6]) + str(1)] = pageFaults
    for i in range(amountOfFrames):
        sheet[str(ascii_uppercase[i + 1]) + str(1)] = i + 1
    k = 2
    if isinstance(frames[0], list):
        for p in range(len(frames)):
            for m in range(len(frames[p])):
                if frames[p][m] not in frames[p - 1]:
                    sheet[str(ascii_uppercase[m + 1]) + str(k)].fill = openpyxl.styles.PatternFill(start_color='FF0000',
                                                                                                   fill_type='solid')
                sheet[str(ascii_uppercase[m + 1]) + str(k)] = frames[p][m]
                # sheet[str(ascii_uppercase[m + 1]) + str(k)] = pages[p]
            k += 1
    elif isinstance(frames[0], dict):
        for p in range(len(frames)):
            i = 0
            for key in frames[p].keys():

                if key not in frames[p - 1] or p==0:
                    sheet[str(ascii_uppercase[i + 1]) + str(k)].fill = openpyxl.styles.PatternFill(start_color='FF0000',
                                                                                                   fill_type='solid')
                sheet[str(ascii_uppercase[i + 1]) + str(k)] = str(key) + '(' + str(frames[p][key]) + ')'
                i += 1
            k += 1
    file.save(fileName)